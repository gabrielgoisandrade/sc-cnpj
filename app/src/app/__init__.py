import csv
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from time import sleep

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By

logging.basicConfig(
    filename="app.log",
    filemode="a",
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    level=logging.INFO,
)

log = logging.getLogger("log")


def get_driver():
    try:
        options = Options()
        options.add_argument("--headless=new")

        service = Service(executable_path=Path("src/driver/chromedriver.exe").resolve())

        driver = Chrome(options, service)
        driver.set_page_load_timeout(10)

        log.info("Driver found")
        return driver
    except FileNotFoundError as e:
        log.exception(msg=e)
        raise e from e


def find_value_by_cnpj(cnpjs: list[str]) -> list[str]:
    log.info(f"Starting search {len(cnpjs)} CNPJs")

    driver = get_driver()

    index_url = "https://appasp.sefaz.go.gov.br/Sintegra/Consulta/default.html"
    result_page_url = "https://appasp.sefaz.go.gov.br/Sintegra/Consulta/consultar.asp"

    results = []

    try:
        driver.get(index_url)

        checkbox = driver.find_element("id", "rTipoDocCNPJ")
        input = driver.find_element("id", "tCNPJ")
        confirm = driver.find_element("name", "btCGC")

        checkbox.click()

        for index, cnpj in enumerate(cnpjs):
            sleep(1.5)
            log.info(f"Searching {index + 1}/{len(cnpjs)}")

            input.clear()
            input.send_keys(cnpj)

            confirm.click()

            sleep(1.5)

            error_tooltip = driver.find_elements(
                By.CLASS_NAME,
                "zion_rich_validation_box",
            )

            if len(error_tooltip):
                log.error(f"{cnpj} is invalid")

                results.append([cnpj, "CNPJ inválido"])
                input.clear()

                continue

            sleep(1.5)

            span = driver.find_elements(
                By.XPATH, "//span[text()='Regime de Apuração:']/following-sibling::*[1]"
            )

            if result_page_url == driver.current_url and not len(span):
                log.info(f"No results found for {cnpj}")

                results.append([cnpj, " "])
                driver.back()

                continue

            log.info(f"Results found: {len(results)} of {len(cnpjs)}")

            results.append([cnpj, span[0].text])

            driver.back()

        return results
    except (WebDriverException, TimeoutException) as e:
        log.exception(msg=e)
        raise e from e
    finally:
        log.info("Finished")
        driver.quit()


def make_chunk(cnpjs):
    chunk_size = 3
    return [cnpjs[i : i + chunk_size] for i in range(0, len(cnpjs), chunk_size)]


def write(v):
    with open("../output", "w+") as f:
        f.write(f"{v[0]} - {v[1]}")


def main() -> None:
    """try:
        file = r"C:/Users/gabriel.andrade/Documents/cnpjs_export_with_situation.xlsx"
        wb = load_workbook(file)
        sheet = wb.active

        cnpjs = [
            str(value)
            for col in sheet.iter_cols(
                min_row=2, max_row=1000, max_col=1, min_col=1, values_only=True
            )
            for value in col
        ]

        results = find_value_by_cnpj(cnpjs)
        situation_column = column_index_from_string("E")
        sheet.cell(column=situation_column, row=1, value="Situação")

        for index, result in enumerate(results, start=2):
            sheet.cell(column=situation_column, row=index, value=result)

        wb.save(r"C:/Users/gabriel.andrade/Documents/cnpjs_export_with_situation.xlsx")
    except (WebDriverException, UnboundLocalError) as e:
        log.exception(msg=e)"""

    with ThreadPoolExecutor(max_workers=2) as executor:
        nums = [
            "45.543.915/0001-81",
            "75.315.333/0001-09",
            "93.209.765/0001-17",
            "03.995.515/0013-09",
        ]

        chunks = make_chunk(nums)

        futures = [executor.submit(find_value_by_cnpj, num) for num in chunks]

        for fut in as_completed(futures):
            [write(item) for item in fut.result()]
